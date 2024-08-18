import uuid
import pandas as pd
from django.http import HttpResponse
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import generics, permissions, status, serializers
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet
from rest_framework.mixins import RetrieveModelMixin, ListModelMixin
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.pagination import PageNumberPagination
from rest_framework.filters import OrderingFilter
from rest_framework.generics import ListAPIView
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import User
from django.db.models import Count
from .models import OrderModel, Comment, Group, STATUS_CHOICES
from .serializers import OrderSerializer, CustomTokenObtainPairSerializer, CommentSerializer, GroupSerializer, \
    UserSerializer, LogoutSerializer, ActivateUserSerializer
from .filters import OrderFilter
from .permissions import IsAdminUserRole


class IsOrderManagerOrReadOnly(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return True
        return obj.manager is None or obj.manager == request.user


class CommentCreateView(generics.CreateAPIView):
    serializer_class = CommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        order = serializer.validated_data['order']
        if order.manager is None or order.manager == self.request.user:
            order.manager = self.request.user
            if order.status not in STATUS_CHOICES or order.status in [None, 'New']:
                order.status = 'In work'
            order.save()
            serializer.save(user=self.request.user)
        else:
            raise serializers.ValidationError('You cannot comment on this order.')


class OrderPagination(PageNumberPagination):
    page_size = 25


class OrderModelViewSet(RetrieveModelMixin, ListModelMixin, GenericViewSet):
    queryset = OrderModel.objects.all().order_by('-id')
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrderManagerOrReadOnly]
    pagination_class = OrderPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = OrderFilter
    ordering_fields = '__all__'
    ordering = ['-id']

    @swagger_auto_schema(
        operation_description="Retrieve a list of orders",
        manual_parameters=[
            openapi.Parameter(
                'course', openapi.IN_QUERY, description="Available values : FS, QACX, JCX, JSCX, FE, PCX",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'course_format', openapi.IN_QUERY, description="Available values : static, online",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'course_type', openapi.IN_QUERY, description="Available values : pro, minimal, premium, incubator, vip",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'status', openapi.IN_QUERY, description="Available values : In work, New, Agree, Disagree, Dubbing",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'ordering', openapi.IN_QUERY, description="Available values : name, -name, id, -id, surname,"
                                                          " -surname, email, -email, phone, -phone, age, -age, course,"
                                                          " -course, course_type, -course_type, course_format,"
                                                          " -course_format, sum, -sum, status, -status, alreadyPaid,"
                                                          " -alreadyPaid, group, -group, created_at, -created_at,"
                                                          " manager, -manager",
                type=openapi.TYPE_STRING),
        ]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Partially update an order",
        request_body=OrderSerializer,
        responses={200: OrderSerializer()},
    )
    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def perform_update(self, serializer):
        serializer.save()


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class CommentListView(ListAPIView):
    serializer_class = CommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        order_id = self.kwargs['order_id']
        return Comment.objects.filter(order_id=order_id).order_by('-created_at')

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        order_id = self.kwargs['order_id']

        try:
            order = OrderModel.objects.get(id=order_id)
        except OrderModel.DoesNotExist:
            return Response({"detail": "Order not found"}, status=404)

        response_data = {
            "comments": serializer.data,
            "utm": order.utm,
            "msg": order.msg,
        }

        return Response(response_data)


class GroupCreateView(generics.CreateAPIView):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [permissions.IsAuthenticated]


class GroupListView(generics.ListAPIView):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [permissions.IsAuthenticated]


class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = LogoutSerializer

    @swagger_auto_schema(
        request_body=LogoutSerializer,
        responses={205: 'Token successfully blacklisted', 400: 'Bad Request'}
    )
    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            refresh_token = serializer.validated_data["refresh"]
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response(status=status.HTTP_205_RESET_CONTENT)
        except Exception as e:
            return Response(status=status.HTTP_400_BAD_REQUEST)


class OrderStatisticsView(APIView):
    permission_classes = [IsAdminUserRole]

    @staticmethod
    def get(request):
        total_count = OrderModel.objects.count()
        statuses = OrderModel.objects.values('status').annotate(count=Count('status'))

        null_count = OrderModel.objects.filter(status__isnull=True).count()

        status_counts = {status['status']: status['count'] for status in statuses}
        status_counts[None] = null_count

        for status in ['In work', 'New', 'Agree', 'Disagree', 'Dubbing', None]:
            if status not in status_counts:
                status_counts[status] = 0

        sorted_statuses = sorted(status_counts.items(), key=lambda x: (x[0] is not None, str(x[0])))

        result = {
            'total_count': total_count,
            'statuses': [{'status': status if status is not None else 'null', 'count': count} for status, count in
                         sorted_statuses]
        }

        return Response(result)


class UserOrderStatisticsView(APIView):
    permission_classes = [IsAdminUserRole]

    @staticmethod
    def get(request, id):
        user = User.objects.get(pk=id)
        orders = OrderModel.objects.filter(manager=user)

        statuses = orders.values('status').annotate(count=Count('status'))

        result = {
            'total_count': orders.count(),
            'statuses': [{'status': status['status'], 'count': status['count']} for status in statuses]
        }

        return Response(result)


class UserListView(APIView):
    permission_classes = [IsAdminUserRole]

    @swagger_auto_schema(
        operation_description="Get a list of users",
        responses={200: UserSerializer(many=True)}
    )
    def get(self, request):
        users = User.objects.all()
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)


class BanUserView(APIView):
    permission_classes = [IsAdminUserRole]

    @staticmethod
    def patch(request, id):
        try:
            user = User.objects.get(pk=id)
            user.is_active = False
            user.save()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except User.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)


class UnbanUserView(APIView):
    permission_classes = [IsAdminUserRole]

    @staticmethod
    def patch(request, id):
        try:
            user = User.objects.get(pk=id)
            user.is_active = True
            user.save()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except User.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)


class CreateUserView(generics.CreateAPIView):
    permission_classes = [IsAdminUserRole]
    serializer_class = UserSerializer

    def perform_create(self, serializer):
        user = serializer.save(is_active=False)
        profile_data = self.request.data.get('profile')

        if profile_data:
            user.userprofile.first_name = profile_data.get('first_name', '')
            user.userprofile.last_name = profile_data.get('last_name', '')
            user.userprofile.save()

        return user

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = self.perform_create(serializer)
        user_data = {
            "id": user.id,
            "email": user.email,
            "is_active": user.is_active,
            "is_superuser": user.is_superuser,
            "is_staff": user.is_staff,
            "create_at": user.date_joined,
            "last_login": user.last_login,
            "profile": {
                "first_name": user.userprofile.first_name,
                "last_name": user.userprofile.last_name
            } if hasattr(user, 'userprofile') else None
        }
        return Response(user_data, status=status.HTTP_201_CREATED)


class UserActivationTokenView(APIView):
    permission_classes = [IsAdminUserRole]

    @swagger_auto_schema(
        operation_description="Get activation token for a user",
        responses={200: "Token", 404: "User not found"}
    )
    def get(self, request, id, *args, **kwargs):
        try:
            user = User.objects.get(pk=id)
            if not user.is_active:
                refresh_token = RefreshToken.for_user(user)
                return Response({"token": str(refresh_token)}, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "User is already active"}, status=status.HTTP_400_BAD_REQUEST)
        except User.DoesNotExist:
            return Response({"detail": "User not found"}, status=status.HTTP_404_NOT_FOUND)


class ActivateUserView(APIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = ActivateUserSerializer

    @swagger_auto_schema(request_body=ActivateUserSerializer)
    def post(self, request, token, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            refresh_token = RefreshToken(token)
            user_id = refresh_token.get('user_id')

            if not user_id:
                return Response({"detail": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

            user = User.objects.get(id=user_id)

            if user.is_active:
                return Response({"detail": "User is already active"}, status=status.HTTP_400_BAD_REQUEST)

            user.set_password(serializer.validated_data["password"])
            user.is_active = True
            user.save()

            return Response({"detail": "User activated successfully"}, status=status.HTTP_200_OK)
        except TokenError:
            return Response({"detail": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)
        except User.DoesNotExist:
            return Response({"detail": "User not found"}, status=status.HTTP_404_NOT_FOUND)


class OrderExcelExportView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = OrderFilter
    ordering_fields = '__all__'
    ordering = ['id']

    @swagger_auto_schema(
        operation_description="Export orders as Excel file",
        responses={200: 'Excel file with orders'},
        manual_parameters=[
            openapi.Parameter(
                'course', openapi.IN_QUERY, description="Available values : FS, QACX, JCX, JSCX, FE, PCX",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'course_format', openapi.IN_QUERY, description="Available values : static, online",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'course_type', openapi.IN_QUERY, description="Available values : pro, minimal, premium, incubator, vip",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'status', openapi.IN_QUERY, description="Available values : In work, New, Agree, Disagree, Dubbing",
                type=openapi.TYPE_STRING),
            openapi.Parameter(
                'ordering', openapi.IN_QUERY, description="Available values : name, -name, id, -id, surname,"
                                                          " -surname, email, -email, phone, -phone, age, -age, course,"
                                                          " -course, course_type, -course_type, course_format,"
                                                          " -course_format, sum, -sum, status, -status, alreadyPaid,"
                                                          " -alreadyPaid, group, -group, created_at, -created_at,"
                                                          " manager, -manager",
                type=openapi.TYPE_STRING),
        ]
    )
    def get(self, request, *args, **kwargs):
        orders = self.filter_queryset(OrderModel.objects.all())

        data = []
        for order in orders:
            created_at = order.created_at
            if created_at is not None and created_at.tzinfo is not None:
                created_at = created_at.replace(tzinfo=None)

            data.append({
                'ID': order.id,
                'Name': order.name,
                'Surname': order.surname,
                'Email': order.email,
                'Phone': order.phone,
                'Age': order.age,
                'Course': order.course,
                'Course Format': order.course_format,
                'Course Type': order.course_type,
                'Status': order.status,
                'Sum': order.sum,
                'Already Paid': order.alreadyPaid,
                'Group': order.group,
                'Created At': created_at,
                'Manager': order.manager.first_name if order.manager else '',
            })

        df = pd.DataFrame(data)

        wb = Workbook()
        ws = wb.active

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            if column_letter == get_column_letter(df.columns.get_loc('Created At') + 1):
                adjusted_width = max(20, max_length + 2)
            else:
                adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        file_name = f'{uuid.uuid4()}.xlsx'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        wb.save(response)

        return response

    def filter_queryset(self, queryset):
        queryset = self.filterset_class(self.request.GET, queryset=queryset).qs
        ordering = self.request.query_params.get('ordering', None)
        if ordering:
            queryset = queryset.order_by(*ordering.split(','))
        return queryset
