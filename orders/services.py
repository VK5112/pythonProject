import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response

from .serializers import OrderSerializer


def update_order_service(instance, validated_data):
    serializer = OrderSerializer(instance, data=validated_data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return serializer.data


def export_orders_to_excel_service(queryset):
    data = []
    for order in queryset:
        created_at = order.created_at
        if created_at is not None and created_at.tzinfo is not None:
            created_at = created_at.replace(tzinfo=None)

        data.append({
            'ID': order.id,
            'Name': order.name,
            'Surname': order.surname,
            'Email': order.email,
            'Phone': order.phone,
            'Age': order.age,
            'Course': order.course,
            'Course Format': order.course_format,
            'Course Type': order.course_type,
            'Status': order.status,
            'Sum': order.sum,
            'Already Paid': order.alreadyPaid,
            'Group': order.group,
            'Created At': created_at,
            'Manager': order.manager.first_name if order.manager else '',
        })

    df = pd.DataFrame(data)

    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                continue

        if column_letter == get_column_letter(df.columns.get_loc('Created At') + 1):
            adjusted_width = max(20, max_length + 2)
        else:
            adjusted_width = max_length + 2

        ws.column_dimensions[column_letter].width = adjusted_width

    file_name = f'{uuid.uuid4()}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    wb.save(response)

    return response


def handle_partial_update_order(instance, data):
    updated_data = update_order_service(instance, data)
    return Response(updated_data, status=status.HTTP_200_OK)
